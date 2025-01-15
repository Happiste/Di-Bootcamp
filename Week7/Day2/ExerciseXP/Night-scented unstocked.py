from openpyxl import load_workbook

workbook_path = 'Plants.xlsx' 
workbook = load_workbook(workbook_path)
sheet = workbook.active 

row = 2  
plants_not_in_stock = []

while True:
    plant_name = sheet.cell(row=row, column=1).value
    stock_status = sheet.cell(row=row, column=8).value 
    if plant_name is None:
        break
    if stock_status == "No":
        plants_not_in_stock.append(plant_name)
    row += 1

print("Plantes qui ne sont pas en stock :")
for plant in plants_not_in_stock:
    print(plant)
