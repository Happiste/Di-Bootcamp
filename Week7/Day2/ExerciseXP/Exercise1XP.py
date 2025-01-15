from openpyxl import Workbook

# Create a new workbook and access the active worksheet
workbook = Workbook()
worksheet = workbook.active

# Add labels and values to the worksheet
worksheet.cell(row=1, column=1, value="First number ==>")
worksheet.cell(row=2, column=1, value="Second number ==>")
worksheet.cell(row=1, column=2, value=2)  # First number
worksheet.cell(row=2, column=2, value=8)  # Second number

# Add a formula to calculate the product of the two numbers
worksheet.cell(row=3, column=2, value="=B1*B2")

# Save the workbook to a file
file_path = 'Simple_Calculator.xlsx' 
workbook.save(file_path)

# Close the workbook
workbook.close()
