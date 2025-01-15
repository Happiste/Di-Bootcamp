from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


data = {
    "Joe": {"math": 65, "science": 78, "english": 98, "gym": 89},
    "Bill": {"math": 55, "science": 72, "english": 87, "gym": 95},
    "Tim": {"math": 100, "science": 45, "english": 75, "gym": 92},
    "Sally": {"math": 30, "science": 25, "english": 45, "gym": 100},
    "Jane": {"math": 100, "science": 100, "english": 100, "gym": 60},
}


workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Grades"


headers = ["Student Names", "Math", "Science", "English", "Gym"]
for col, header in enumerate(headers, start=1):
    cell = worksheet.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")


for row, (student, grades) in enumerate(data.items(), start=2):
    worksheet.cell(row=row, column=1, value=student) 
    for col, subject in enumerate(grades, start=2):
        worksheet.cell(row=row, column=col, value=grades[subject])


last_row = len(data) + 2  
for col in range(2, len(headers) + 1):  
    column_letter = worksheet.cell(row=1, column=col).column_letter
    worksheet.cell(row=last_row, column=col, value=f"=AVERAGE({column_letter}2:{column_letter}{last_row - 1})")


file_path = "Grades.xlsx"
workbook.save(file_path)

print(f"The workbook has been saved as {file_path}")
